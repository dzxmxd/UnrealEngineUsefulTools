import re
import pandas as pd
import sys
import os
from openpyxl import load_workbook


def parse_log_to_excel(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        log_data = f.read()

    widget_data = {}
    current_widget = None

    for line in log_data.split("\n"):
        widget_match = re.match(r'.*====\s+(\w+_C):\s+(\d+)\s+====', line)
        component_match = re.match(r'.*\s+(\w+):\s+(\d+)', line)

        if widget_match:
            current_widget = widget_match.group(1)
            widget_data[current_widget] = {"Instance Count": int(widget_match.group(2))}
        elif component_match and current_widget:
            widget_data[current_widget][component_match.group(1)] = int(component_match.group(2))

    all_columns = set()
    for data in widget_data.values():
        all_columns.update(data.keys())

    all_columns = ["UserWidget"] + sorted(all_columns - {"Instance Count"})

    rows = []
    for widget, data in widget_data.items():
        row = {col: data.get(col, "") for col in all_columns}  # 为空时使用空字符串
        row["UserWidget"] = widget
        row["Instance Count"] = data.get("Instance Count", "")
        rows.append(row)

    df = pd.DataFrame(rows, columns=["UserWidget", "Instance Count"] + sorted(all_columns[1:]))

    output_path = os.path.splitext(file_path)[0] + ".xlsx"
    df.to_excel(output_path, index=False, engine='openpyxl')

    # 确保数字以数值格式存储
    wb = load_workbook(output_path)
    ws = wb.active
    for col in range(2, ws.max_column + 1):  # 从第二列开始
        for row in range(2, ws.max_row + 1):  # 从第二行开始
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.isdigit():
                cell.value = int(cell.value)
    wb.save(output_path)
    print(f"Excel 文件已生成: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("请提供日志文件路径作为参数")
        sys.exit(1)
    parse_log_to_excel(sys.argv[1])


"""
void FLyraObjectCC::PrintUserWidgetChildrenNumber(const TArray<FString>& Args)
{
    // 用于存储每个 UserWidget 类的统计信息
    TMap<FString, int32> WidgetClassCount; // 每种 UserWidget 类型的对象数量
    TMap<FString, TMap<FString, int32>> WidgetClassChildStats; // 每种 UserWidget 类型的子控件统计

    // 遍历所有 UUserWidget 对象
    for (TObjectIterator<UUserWidget> It; It; ++It)
    {
        const UUserWidget* UserWidget = *It;
        if (UserWidget && UserWidget->WidgetTree) // 确保 WidgetTree 有效
        {
            // 获取 UserWidget 的类名
            FString WidgetClassName = UserWidget->GetClass()->GetName();

            // 更新该类名的对象数量
            if (WidgetClassCount.Contains(WidgetClassName))
            {
                WidgetClassCount[WidgetClassName]++;
            }
            else
            {
                WidgetClassCount.Add(WidgetClassName, 1);
                WidgetClassChildStats.Add(WidgetClassName, TMap<FString, int32>());
            }

            // 获取 UserWidget 的所有子控件
            TArray<UWidget*> Children;
            UserWidget->WidgetTree->GetAllWidgets(Children);

            // 统计每个子控件的类型
            for (UWidget* Child : Children)
            {
                if (Child)
                {
                    FString ChildClassName = Child->GetClass()->GetName();

                    // 更新该类名的子控件统计
                    if (WidgetClassChildStats[WidgetClassName].Contains(ChildClassName))
                    {
                        WidgetClassChildStats[WidgetClassName][ChildClassName]++;
                    }
                    else
                    {
                        WidgetClassChildStats[WidgetClassName].Add(ChildClassName, 1);
                    }
                }
            }
        }
    }

    // 打印统计结果
    for (const auto& WidgetClassPair : WidgetClassCount)
    {
        FString WidgetClassName = WidgetClassPair.Key;
        const int32 WidgetCount = WidgetClassPair.Value;

        // 打印 UserWidget 类名及其对象数量
        UE_LOG(LogTemp, Log, TEXT("==== %s: %d ===="), *WidgetClassName, WidgetCount);

        // 打印每个子控件类型的数量
        if (WidgetClassChildStats.Contains(WidgetClassName))
        {
            const TMap<FString, int32>& ChildStats = WidgetClassChildStats[WidgetClassName];
            for (const auto& ChildStat : ChildStats)
            {
                UE_LOG(LogTemp, Log, TEXT("    %s: %d"), *ChildStat.Key, ChildStat.Value);
            }
        }
    }
}
"""